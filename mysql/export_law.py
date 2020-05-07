#!/usr/bin/env python
# _*_ coding:utf-8 _*_
#
# @Version : 1.0
# @Time    : 2018-06-12 12:03:04
# @Author  : liupan
# @File    : export_law.py
#
# 写数据到Excel

import os
from openpyxl import load_workbook

# 获取附件目录文件列表，如果没有在urlResultList里面，就删除
# def GetFileList(dir, fileList):
#     newDir = dir
#     if os.path.isfile(dir):
#         fileList.append(dir)
#     elif os.path.isdir(dir):
#         for s in os.listdir(dir):
#             # 如果需要忽略某些文件夹，使用以下代码
#             # if s == "xxx":
#             # continue
#             newDir = os.path.join(dir, s)
#             GetFileList(newDir, fileList)
#     return fileList
# fileList = GetFileList('/root/fileUpload/merchant/attachment', [])

# 写数据
newfile = './ret.xlsx'
wb = load_workbook(newfile)
ws = wb['Sheet1']
#直接根据位置进行赋值
ws['A1'] = '类别'
ws['B1'] = '法规名称'
ws['C1'] = '是否存在'


dir = '/Users/panliu/Downloads/1'
i = 2
if os.path.isdir(dir):
    for s in os.listdir(dir):
        print(s)
        ws.cell(row=i, column=1, value=str(s))
        newDir = os.path.join(dir, s)
        if os.path.isdir(newDir):
            for s2 in os.listdir(newDir):
                ws.cell(row=i, column=2, value=str(s2))
                i = i + 1

wb.save(newfile)
