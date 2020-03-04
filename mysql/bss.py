#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2020/3/4 上午11:20
# @Author  : liupan
# @Site    : 
# @File    : bss.py
# @Software: PyCharm

import openpyxl
from mysql_util import MysqlUtil


def read_excel():
    excel_file = r'./bss.xlsx'
    inwb = openpyxl.load_workbook(excel_file)
    ws = inwb['Sheet1']

    # 获取sheet的最大行数和列数
    rows = ws.max_row
    cols = ws.max_column
    all_item = []
    for r in range(2, rows + 1):
        id = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        seq = ws.cell(r, 4).value
        area_id = ws.cell(r, 5).value
        pid = ws.cell(r, 6).value
        item = {
            'id': id,
            'name': name,
            'code': code,
            'seq': seq,
            'area_id': area_id,
            'pid': pid
        }
        all_item.append(item)
    return all_item


if __name__ == '__main__':
    all_item = read_excel()
    sql = MysqlUtil()
    for item in all_item:
        sql.insert('t_gov_dept', **item)
